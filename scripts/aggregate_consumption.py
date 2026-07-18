#!/usr/bin/env python3
"""Generate reproducible visitor-consumption aggregates from the official workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import statistics
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CATEGORY_FIELDS = {
    "ticket": "ticket_cost",
    "food": "food_cost",
    "shopping": "shopping_cost",
    "transport": "transport_cost",
    "entertainment": "entertainment_cost",
}
NUMERIC_FIELDS = [
    *CATEGORY_FIELDS.values(),
    "total_cost",
    "age",
    "stay_duration",
    "group_size",
    "satisfaction",
]
AGE_GROUPS = [
    (0, 17, "0-17"),
    (18, 25, "18-25"),
    (26, 35, "26-35"),
    (36, 45, "36-45"),
    (46, 60, "46-60"),
    (61, math.inf, "61+"),
]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def round_money(value: float) -> float:
    return round(value + 1e-9, 2)


def age_group(age: float | None) -> str:
    if age is None:
        return "unknown"
    for lower, upper, label in AGE_GROUPS:
        if lower <= age <= upper:
            return label
    return "unknown"


def month_value(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    text = str(value or "").strip()
    if len(text) >= 7 and text[4] == "-":
        return text[:7]
    return "unknown"


def average(values: list[float]) -> float:
    return round_money(statistics.fmean(values)) if values else 0.0


def build_group(rows: list[dict[str, Any]]) -> dict[str, Any]:
    costs = [row["total_cost"] for row in rows if row["total_cost"] is not None]
    return {
        "sample_count": len(rows),
        "total_cost": round_money(sum(costs)),
        "average_total_cost": average(costs),
        "median_total_cost": round_money(statistics.median(costs)) if costs else 0.0,
        "average_stay_duration": average(
            [row["stay_duration"] for row in rows if row["stay_duration"] is not None]
        ),
        "average_group_size": average(
            [row["group_size"] for row in rows if row["group_size"] is not None]
        ),
        "average_satisfaction": average(
            [row["satisfaction"] for row in rows if row["satisfaction"] is not None]
        ),
    }


def aggregate(
    rows: list[dict[str, Any]],
    scope: str,
    source_metadata: dict[str, Any],
    quality: dict[str, Any],
) -> dict[str, Any]:
    total_costs = [row["total_cost"] for row in rows if row["total_cost"] is not None]
    category_totals: dict[str, float] = {}
    category_counts: dict[str, int] = {}
    for category, field in CATEGORY_FIELDS.items():
        values = [row[field] for row in rows if row[field] is not None]
        category_totals[category] = round_money(sum(values))
        category_counts[category] = len(values)

    category_total = sum(category_totals.values())
    category_breakdown = [
        {
            "category": category,
            "sample_count": category_counts[category],
            "total_cost": category_totals[category],
            "share_percent": round(category_totals[category] / category_total * 100, 2)
            if category_total
            else 0.0,
        }
        for category in CATEGORY_FIELDS
    ]

    monthly_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    age_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    satisfaction_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    stay_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    group_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        monthly_rows[month_value(row["visit_date"])].append(row)
        age_rows[age_group(row["age"])].append(row)
        satisfaction = row["satisfaction"]
        satisfaction_rows[
            str(int(satisfaction))
            if satisfaction is not None and satisfaction.is_integer()
            else "unknown"
        ].append(row)
        stay = row["stay_duration"]
        stay_label = (
            "unknown"
            if stay is None
            else "0-2h"
            if stay < 2
            else "2-4h"
            if stay < 4
            else "4-8h"
            if stay < 8
            else "8h+"
        )
        stay_rows[stay_label].append(row)
        group_size = row["group_size"]
        group_label = (
            "unknown"
            if group_size is None
            else "1"
            if group_size <= 1
            else "2"
            if group_size == 2
            else "3-4"
            if group_size <= 4
            else "5+"
        )
        group_rows[group_label].append(row)

    monthly_trend = [
        {"month": month, **build_group(monthly_rows[month])}
        for month in sorted(monthly_rows)
    ]
    segments = {
        "age_groups": [
            {"segment": key, **build_group(age_rows[key])} for key in sorted(age_rows)
        ],
        "satisfaction": [
            {"segment": key, **build_group(satisfaction_rows[key])}
            for key in sorted(satisfaction_rows)
        ],
        "stay_duration": [
            {"segment": key, **build_group(stay_rows[key])} for key in sorted(stay_rows)
        ],
        "group_size": [
            {"segment": key, **build_group(group_rows[key])}
            for key in sorted(group_rows)
        ],
    }

    recommendations: list[str] = []
    if category_total:
        top_category = max(category_totals, key=category_totals.get)
        recommendations.append(
            f"消费结构中{top_category}占比最高，建议优先检查该类别的服务承载和转化路径。"
        )
    if len(total_costs) >= 2 and statistics.fmean(total_costs) > 0:
        recommendations.append(
            "可结合停留时长、同行人数和年龄段进一步设计分层套餐，不把相关性直接解释为因果关系。"
        )
    if not recommendations:
        recommendations.append(
            "当前样本不足以形成稳定运营建议，建议补充数据后再观察趋势。"
        )

    unique_tourists = len({row["tourist_id"] for row in rows if row["tourist_id"]})
    unique_attractions = len(
        {row["attraction_name"] for row in rows if row["attraction_name"]}
    )
    return {
        "scope": scope,
        "source_metadata": source_metadata,
        "summary": {
            "sample_count": len(rows),
            "unique_tourists": unique_tourists,
            "unique_attractions": unique_attractions,
            "total_cost": round_money(sum(total_costs)),
            "average_total_cost": average(total_costs),
            "median_total_cost": round_money(statistics.median(total_costs))
            if total_costs
            else 0.0,
            "average_stay_duration": average(
                [
                    row["stay_duration"]
                    for row in rows
                    if row["stay_duration"] is not None
                ]
            ),
            "average_group_size": average(
                [row["group_size"] for row in rows if row["group_size"] is not None]
            ),
            "average_satisfaction": average(
                [row["satisfaction"] for row in rows if row["satisfaction"] is not None]
            ),
        },
        "category_breakdown": category_breakdown,
        "monthly_trend": monthly_trend,
        "segments": segments,
        "recommendations": recommendations,
        "data_quality": quality,
    }


def read_rows(
    input_path: Path, name_filter: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(iterator)]
    required = {"tourist_id", "attraction_name", "visit_date", *NUMERIC_FIELDS}
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"missing workbook columns: {', '.join(missing)}")

    index = {name: position for position, name in enumerate(headers)}
    all_rows: list[dict[str, Any]] = []
    filtered_rows: list[dict[str, Any]] = []
    quality = {
        "total_rows": 0,
        "invalid_numeric_rows": 0,
        "invalid_values_by_field": defaultdict(int),
        "missing_values_by_field": defaultdict(int),
        "negative_values_by_field": defaultdict(int),
        "derived_total_cost_count": 0,
    }
    for values in iterator:
        quality["total_rows"] += 1
        row = {
            name: values[position] if position < len(values) else None
            for name, position in index.items()
        }
        invalid_row = False
        for field in NUMERIC_FIELDS:
            original = row[field]
            number = as_number(original)
            if original in (None, ""):
                quality["missing_values_by_field"][field] += 1
            elif number is None:
                quality["invalid_values_by_field"][field] += 1
                invalid_row = True
            elif number < 0 and field in [
                *CATEGORY_FIELDS.values(),
                "total_cost",
                "stay_duration",
                "group_size",
            ]:
                quality["negative_values_by_field"][field] += 1
                invalid_row = True
            row[field] = number
        if row["total_cost"] is None:
            category_values = [row[field] for field in CATEGORY_FIELDS.values()]
            if all(value is not None and value >= 0 for value in category_values):
                row["total_cost"] = sum(category_values)
                quality["derived_total_cost_count"] += 1
        if invalid_row:
            quality["invalid_numeric_rows"] += 1
        all_rows.append(row)
        search_text = (
            f"{row.get('attraction_name') or ''} {row.get('attraction_content') or ''}"
        )
        if name_filter in search_text:
            filtered_rows.append(row)

    workbook.close()
    quality = {
        key: dict(value) if isinstance(value, defaultdict) else value
        for key, value in quality.items()
    }
    return all_rows, filtered_rows, {"headers": headers, "quality": quality}


def build_output(input_path: Path, name_filter: str) -> dict[str, Any]:
    all_rows, filtered_rows, info = read_rows(input_path, name_filter)
    source_metadata = {
        "source_file": input_path.name,
        "source_sha256": sha256_file(input_path),
        "source_row_count": len(all_rows),
        "source_columns": info["headers"],
        "name_filter": name_filter,
        "generated_at": datetime.now().astimezone().isoformat(),
    }
    quality = info["quality"]
    return {
        "schema_version": 1,
        "source_metadata": source_metadata,
        "all": aggregate(all_rows, "all", source_metadata, quality),
        "lingshan": aggregate(filtered_rows, "lingshan", source_metadata, quality),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--name-filter", default="灵山")
    args = parser.parse_args()
    if not args.input.is_file():
        raise SystemExit(f"input file not found: {args.input}")
    output = build_output(args.input, args.name_filter)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "output": str(args.output),
                "source_sha256": output["source_metadata"]["source_sha256"],
                "all_rows": output["source_metadata"]["source_row_count"],
                "lingshan_rows": output["lingshan"]["summary"]["sample_count"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
